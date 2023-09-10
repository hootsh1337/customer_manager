# -*- coding: utf-8 -*-
"""
Created on Wed Aug 23 01:51:27 2023

@author: H
"""

# IMPORTS
import tkinter as tk
from tkinter import ttk
import openpyxl
import os
import pywhatkit
import pandas as pd
import re
# END IMPORTS

# This project attempts to create a small customer relationship management system
# The two main feature at first are:
# Easily view, edit, add and remove clients and their info from a simple database, using a minimal GUI
# Allow the population of the client data by importing Excel sheets
# Add an option for sending WhatsApp messages and notifications to customers

# LOG START
# -23AUG--CREATED GUI INFRASTRUCTURE--(NEED TO REVIEW CONCEPTS)
# -28AUG--ADDED BUTTONS AND ENTRY BOXES--NOT YET FUNCTIONAL
# -30AUG--UPDATED SOME BUTTONS FUNCTIONALITY
# -08SEP--INTRODUCED EXCEL IMPORT FUNCTIONALITY
# -09SEP--ADD RECORD BUTTON IS FUNCTIONAL
# -10SEP--INTRODUCED REFRESH FUNCTIONALITY--UPDATE AND SINGLE ENTRY DELETE NOW AFFECTS XL FILE--INTRODUCED ENTRY COUNTER
# -TODO---FIX WHATSAPP BUTTON + UPDATE DELETE BUTTONS
# LOG END

total_clients = 0

#############################################################################

# Load data to viewer

def load_data():
    current_file_path = os.path.dirname(__file__)
    xl_file_path = current_file_path + "\customers.xlsx"
    workbook = openpyxl.load_workbook(xl_file_path)
    sheet = workbook.active
    
    total_clients = sheet.max_row - 1
    total_clients = str(total_clients)
    print("total number of clients =",total_clients)
    
    list_values = list(sheet.values)

    for value_tuple in list_values[1:]:
        my_tree.insert('', tk.END, values=value_tuple)
    
    workbook.close()
    
    
    total_number_value = ttk.Label(data_frame, text=total_clients)
    total_number_value.grid(row=0, column=9, padx=10, pady=10)

#############################################################################

# Remove record functionality


def remove_one():
    
    # grab record number
    selected = my_tree.focus()
    # grab client id number from record
    values = my_tree.item(selected, 'values')
    id_number= int(values[0])
    
    print("Deleted entry number:", id_number)
    
    # open sheet 
    current_file_path = os.path.dirname(__file__)
    xl_file_path = current_file_path + "\customers.xlsx"
    df = pd.read_excel(xl_file_path)
    
    # find selected client id in sheet and drop
    df2=df.drop(df[df['id'] == id_number].index)
    #print(df[df['id'] == id_number].index)
    #print(df2)

    # save
    with pd.ExcelWriter(xl_file_path) as writer:
        df2.to_excel(writer, sheet_name= "Sheet", index=False)
    
    # clear treeview to refresh
    for item in my_tree.get_children():
        my_tree.delete(item)
   
    # load data again
    load_data()


def remove_all():
    for record in my_tree.get_children():
        my_tree.delete(record)


#############################################################################
# Update record functionality


def update_record():
    # grab rec number
    selected = my_tree.focus()

    # grab client id number from record
    values = my_tree.item(selected, 'values')
    id_number= int(values[0])

    print("Updated client ID:", id_number)
    
    # update record in program
    # my_tree.item(selected, text="", values=(name_entry.get(), adrs_entry.get(), area_entry.get(), ph_entry.get(), ph2_entry.get(), tax_entry.get(), fees_entry.get(), cmts_entry.get(),))
    
    # remember data 
    name = str(name_entry.get())
    address = str(adrs_entry.get())
    area = str(area_entry.get())
    phone = str(ph_entry.get())
    phone2 = str(ph2_entry.get())
    tax = str(tax_entry.get())
    comments = str(cmts_entry.get())
    
    # open sheet add new row, delete old row, and save
    current_file_path = os.path.dirname(__file__)
    xl_file_path = current_file_path + "\customers.xlsx"
    
    df = pd.read_excel(xl_file_path)
    
    # find record in df and update
    df.loc[df['id'] == id_number] = [id_number , name , address , area , phone , phone2 , tax , comments]
    
    print(df)

    # save
    with pd.ExcelWriter(xl_file_path) as writer:
        df.to_excel(writer, sheet_name= "Sheet", index=False)
    
    # clear the entry boxes
    name_entry.delete(0, "end")
    adrs_entry.delete(0, "end")
    area_entry.delete(0, "end")
    ph_entry.delete(0, "end")
    ph2_entry.delete(0, "end")
    tax_entry.delete(0, "end")
    cmts_entry.delete(0, "end")
    
    # clear treeview to refresh
    for item in my_tree.get_children():
        my_tree.delete(item)
   
    # load data again
    load_data()
    
#############################################################################
# Clear boxes functionality


def clear_boxes():

    #clear the entry boxes
    name_entry.delete(0, "end")
    adrs_entry.delete(0, "end")
    area_entry.delete(0, "end")
    ph_entry.delete(0, "end")
    ph2_entry.delete(0, "end")
    tax_entry.delete(0, "end")
    cmts_entry.delete(0, "end")
    
#############################################################################
# Add record functionality


def add_record():

    # get data from boxes
    name = name_entry.get()
    address = adrs_entry.get()
    area = area_entry.get()
    phone = ph_entry.get()
    phone2 = ph2_entry.get()
    tax = tax_entry.get()
    comments = cmts_entry.get()
    
    # open excel sheet
    current_file_path = os.path.dirname(__file__)
    xl_file_path = current_file_path + "\customers.xlsx"
    
    df = pd.read_excel(xl_file_path)
    
    # make new id number
    max_previous_id = df['id'].max()
    print("Last client id=",max_previous_id)
    new_record_id = max_previous_id + 1
    print("Newest client id=",new_record_id)
    
    row_values = [new_record_id,name,address,area,phone,phone2,tax,comments]
    
    df2=df.append(row_values)
    
    
    
    # insert new row in treeview
    my_tree.insert('', tk.END, values=row_values)
    
    #clear the entry boxes
    name_entry.delete(0, "end")
    adrs_entry.delete(0, "end")
    area_entry.delete(0, "end")
    ph_entry.delete(0, "end")
    ph2_entry.delete(0, "end")
    tax_entry.delete(0, "end")
    cmts_entry.delete(0, "end")
    
#############################################################################
    
def send_wapp():
    
    current_file_path = os.path.dirname(__file__)
    xl_file_path = current_file_path + "\customers.xlsx"
    workbook = openpyxl.load_workbook(xl_file_path)
    sheet = workbook.active
    
    list_values = list(sheet.values)

    for value_tuple in list_values[1:]:
        my_tree.insert('', tk.END, values=value_tuple)
    
        if tax_entry == 'Late Tax Statement':
            # Send a whatsapp message at 6PM, message is delivered after 60 seconds
            pywhatkit.sendwhatmsg("","",18,00,60)


#############################################################################

# Style
root = tk.Tk()
root.title = ('CRM Test')
root.geometry = ("1000x500")
root.iconbitmap = ('F:\Art\Icons\pc.ico')
style = ttk.Style(root)

style.theme_use('default')

# Treeview colors
style.configure("Treeview", background="#D3D3D3", foreground="black", rowheight=25, fieldbackground="#D3D3D3")

# Change color of selected client
style.map('Treeview', background=[('selected', "347083")])

# Frame
frame = ttk.Frame(root)
frame.pack(pady=10)



#############################################################################
# Add record entry boxes
#############################################################################

# Frame
data_frame = ttk.LabelFrame(root, text="Record")
data_frame.pack(fill="x", expand="yes", padx=20)

####################################################

name_label = ttk.Label(data_frame, text="Name")
name_label.grid(row=0, column=0, padx=10, pady=10)

name_entry = ttk.Entry(data_frame)
name_entry.grid(row=0, column=1, padx=10, pady=10)
####################################################
adrs_label = ttk.Label(data_frame, text="Address")
adrs_label.grid(row=0, column=2, padx=10, pady=10)

adrs_entry = ttk.Entry(data_frame)
adrs_entry.grid(row=0, column=3, padx=10, pady=10)
####################################################
area_label = ttk.Label(data_frame, text="Area")
area_label.grid(row=0, column=4, padx=10, pady=10)

area_entry = ttk.Entry(data_frame)
area_entry.grid(row=0, column=5, padx=10, pady=10)
####################################################
ph_label = ttk.Label(data_frame, text="Phone")
ph_label.grid(row=1, column=0, padx=10, pady=10)

ph_entry = ttk.Entry(data_frame)
ph_entry.grid(row=1, column=1, padx=10, pady=10)
####################################################
ph2_label = ttk.Label(data_frame, text="Phone 2")
ph2_label.grid(row=1, column=2, padx=10, pady=10)

ph2_entry = ttk.Entry(data_frame)
ph2_entry.grid(row=1, column=3, padx=10, pady=10)
####################################################
# tax_label = ttk.Label(data_frame, text="Tax Notes")
# tax_label.grid(row=1, column=4, padx=10, pady=10)
# 
# tax_entry = ttk.Entry(data_frame)
# tax_entry.grid(row=1, column=5, padx=10, pady=10)
####################################################
id_label = ttk.Label(data_frame, text="Client ID")
id_label.grid(row=1, column=6, padx=10, pady=10)


####################################################
cmts_label = ttk.Label(data_frame, text="General Notes")
cmts_label.grid(row=1, column=4, padx=10, pady=10)

cmts_entry = ttk.Entry(data_frame)
cmts_entry.grid(row=1, column=5, padx=10, pady=10)
####################################################
status_label = ttk.Label(data_frame, text="Tax Status")
status_label.grid(row=0, column=6, padx=10, pady=10)

status_list = ["Late Tax Statement", "Done", "On Hold"]
tax_entry = ttk.Combobox(data_frame, values=status_list)
tax_entry.current(0)
tax_entry.grid(row=0, column=7, padx=5, pady=5,  sticky="ew")
####################################################
total_number_label = ttk.Label(data_frame, text="Total Clients")
total_number_label.grid(row=0, column=8, padx=10, pady=10)

total_number_value = ttk.Label(data_frame, text=total_clients)
total_number_value.grid(row=0, column=9, padx=10, pady=10)


#####################################################################
# Add buttons
#####################################################################

# Frame
button_frame = ttk.LabelFrame(root, text="Commands")
button_frame.pack(fill="x", expand="yes", padx=20)

####################################################

add_button = ttk.Button(button_frame, text="Add New Client", command=add_record)
add_button.grid(row=0, column=0, padx=10, pady=10)

upd_button = ttk.Button(button_frame, text="Update Client Details", command=update_record)
upd_button.grid(row=0, column=1, padx=10, pady=10)

rmv_button= ttk.Button(button_frame, text="Remove Selected Client Details", command=remove_one)
rmv_button.grid(row=0, column=2, padx=10, pady=10)

rmvall_button= ttk.Button(button_frame, text="Remove All Clients Details", command=remove_all)
rmvall_button.grid(row=0, column=3, padx=10, pady=10)

clr_button= ttk.Button(button_frame, text="Clear", command=clear_boxes)
clr_button.grid(row=0, column=4, padx=10, pady=10)

wapp_button= ttk.Button(button_frame, text="Send WhatsApp Reminder to Late Clients")
wapp_button.grid(row=0, column=5, padx=10, pady=10)

#####################################################################
# TreeView Frame
#####################################################################

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

#####################################################################
# Create the treeview
#####################################################################

my_tree = ttk.Treeview(treeFrame, yscrollcommand=treeScroll.set, selectmode="extended")

#Creating Columns
my_tree['columns'] = ("ID", "Name", "Address", "Area", "Phone", "Phone 2", "Tax Status", "Notes")

my_tree.column("#0", width=0)
my_tree.column("ID", width=140)
my_tree.column("Name", width=140)
my_tree.column("Address", width=140)
my_tree.column("Area", width=140)
my_tree.column("Phone", width=140)
my_tree.column("Phone 2", width=140)
my_tree.column("Tax Status", width=140)

my_tree.column("Notes", width=140)

# Creating Headings
my_tree.heading("#0", text="")
my_tree.heading("ID", text="Client ID")
my_tree.heading("Name", text="Name")
my_tree.heading("Address", text="Address")
my_tree.heading("Area", text="Area")
my_tree.heading("Phone", text="Phone")
my_tree.heading("Phone 2", text="Phone 2")
my_tree.heading("Tax Status", text="Tax Status")
my_tree.heading("Notes", text="Notes")


#############################################################################

# Select Record functionality


def select_record(e):
    # first clear the entry boxes
    name_entry.delete(0, "end")
    adrs_entry.delete(0, "end")
    area_entry.delete(0, "end")
    ph_entry.delete(0, "end")
    ph2_entry.delete(0, "end")
    tax_entry.delete(0, "end")
    cmts_entry.delete(0, "end")

    # grab record number
    selected = my_tree.focus()
    # grab record values
    values = my_tree.item(selected, 'values')

    # output to entry boxes
    id_number= values[0]
    name_entry.insert(0, values[1])
    adrs_entry.insert(0, values[2])
    area_entry.insert(0, values[3])
    ph_entry.insert(0, values[4])
    ph2_entry.insert(0, values[5])
    tax_entry.insert(0, values[6])
    cmts_entry.insert(0, values[7])

    id_value = ttk.Label(data_frame, text=id_number)
    id_value.grid(row=1, column=7, padx=10, pady=10)

############################################################################
# Selected Record Shows in Boxes 
my_tree.bind("<ButtonRelease-1>", select_record)
############################################################################



my_tree.pack()
treeScroll.config(command=my_tree.yview)
load_data()


root.mainloop()
